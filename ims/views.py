import io
from datetime import datetime, timedelta

import pandas as pd
import qrcode
from django.contrib.auth.decorators import login_required
from django.core.files.base import ContentFile
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.utils import timezone
from openpyxl.styles import Alignment
from xhtml2pdf import pisa

from .forms import Assign_form, product_form
from .models import Assign, Category, Department, Product


# Create your views here.
@login_required
def index(request):
    assign = None
    error_message = None
    success_message = None
    product  = Product.objects.all()
    department = Department.objects.all()
    category = Category.objects.all()

    if request.method == 'POST':
        form = Assign_form(request.POST)
        if form.is_valid():
            assign = form.save(commit=False) 
            productassign = assign.product
            if productassign.quantity >= assign.quantity:
                productassign.quantity -= assign.quantity
                productassign.save()
                assign.save()
                success_message = "the product has been assigned." 

        else:
          error_message = "Insufficient quantity available." 

    form = Assign_form()
    context = {'product': product, 'department': department,'category':category  ,'form': form, 'assign': assign, 'error_message': error_message, 'success_message':success_message }
    
    return render(request, 'ims/index.html', context)

@login_required
def assigned_product(request):
    product = Assign.objects.all()
    context = { 'assigned_product':product}
    return render(request, 'ims/products.html', context)

@login_required
def add_product(request):
    success_message = None
    error_message = None
    if request.method == "POST":
       form = product_form(request.POST)
       if form.is_valid():
           form.save()
           success_message = "successfully created!"
       else: 
           error_message = "The product is already added please update!" 
    else:
        
        form = product_form()
        
    context =  {"form": form, 'success_message':success_message, 'error_message':error_message}

    return render(request, "ims/addproduct.html", context)

def product_detail(request, pk):
    product = Product.objects.filter(id = pk)
    context = {'product': product}
    return render(request, "ims/details.html", context)

def department_view(request):
    q = request.GET.get('q') if request.GET.get('q') != None  else ''
    department_view = Assign.objects.filter(department__name=q) 
    context = {'department_view':department_view }

    return render(request, 'ims/view_departments.html', context ) 

def category_view(request):
    q = request.GET.get('q') if request.GET.get('q') != None  else ''
    category_view = Assign.objects.filter(product__product_category__name=q)
    context = {'category_view':category_view }

    return render(request, 'ims/view_category.html', context ) 

#delete store products
def delete(request,pk):
    product = Product.objects.get(id=pk)
    if request.method == 'POST':
        product.delete()
        return redirect('index')
    return render(request, 'ims/delete.html',{'obj':product})

#delete assined products
def delete_product(request,pk):
    product = Assign.objects.get(id=pk)
    if request.method == 'POST':
        product.delete()
        return redirect('product')
    
    return render(request, 'ims/delete.html',{'obj':product})

#delete store products
def update(request, pk):
    success_message = None
    project = Product.objects.get(id=pk)
    form = product_form(instance=project)
    if request.method == 'POST':
        form = product_form(request.POST, instance=project)
        if form.is_valid():
            form.save()
            return redirect('index')
        success_message = "Product successfully Updated!"
    context = {'form': form, 'success_message': success_message}
    
    return render(request, 'ims/addproduct.html', context)

#update assined products
def update_product(request, pk):
    success_message = None
    project = Assign.objects.get(id=pk)
    form = Assign_form(instance=project)
    if request.method == 'POST':
        form =Assign_form(request.POST, instance=project)
        if form.is_valid():
            form.save()
            return redirect('product')
        success_message = "Product successfully Updated!"
    context = {'form': form, 'success_message': success_message}
    
    return render(request, 'ims/assign.html', context)


def generate_invoice(request, pro_id):
    assign = get_object_or_404(Assign, id=pro_id)

    generate_qr_code(assign)
    content = render_to_string('ims/invoice.html', {'assign': assign})

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=invoice_{timezone.now()}.pdf'
    status = pisa.CreatePDF(content, dest=response)

    if status.err:
        return HttpResponse("Error creating PDF", status=500)
    return response


def generate_qr_code(assign):
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(f"Assign ID: {assign.pro_id} \n Name: {assign.name}  \n product: {assign.product}  \n department: {assign.department}"  )
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    img_bytes = io.BytesIO()
    img.save(img_bytes, format='PNG')
    assign.qr_code.save(f'qr_code_{assign.id}.png', ContentFile(img_bytes.getvalue()))


@login_required
def generate_report(request, period):
    if period == 'daily':
        start_date = datetime.now().date()
        end_date = start_date + timedelta(days=1)
    elif period == 'weekly':
        start_date = (datetime.now() - timedelta(days=datetime.now().weekday())).date()
        end_date = start_date + timedelta(weeks=1)
    elif period == 'monthly':
        start_date = datetime.now().replace(day=1).date()
        end_date = (start_date.replace(month=start_date.month % 12 + 1, day=1))
    else:
        return HttpResponse("Invalid period", status=400)

    assign = Assign.objects.filter(assigned_date__range=(start_date, end_date))
    if request.method == 'POST':
        data = {
                'ID': [assign.id for assign in assign],
                'Name': [assign.name for assign in assign],
                'Product': [assign.product for assign in assign],
                'Quantity': [assign.quantity for assign in assign],
                'Date Ordered': [assign.assigned_date for assign in assign],
            }
        df = pd.DataFrame(data)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=sales_report_{period}.xlsx'
        with pd.ExcelWriter(response, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Sales Report', index=False)
                
                workbook = writer.book
                worksheet = writer.sheets['Sales Report']

                for row in worksheet.iter_rows(min_row=1, max_col=worksheet.max_column, max_row=worksheet.max_row):
                    for cell in row:
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                for column in worksheet.columns:
                    max_length = 0
                    column = [col for col in column]
                    try:
                        max_length = max(len(str(cell.value)) for cell in column)
                        adjusted_width = (max_length + 2)
                        worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
                    except:
                        pass

        return response
    else:
        return render(request, 'ims/reports.html')
    
@login_required
def assign_report(request):
    if request.method == 'POST':
        period = request.POST.get('period', 'daily')  
        return generate_report(request, period)
    return render(request, 'ims/reports.html')



